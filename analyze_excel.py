import openpyxl
import sqlite3
import json
from pathlib import Path
import os

def analyze_excel_structure(file_path):
    """Analyze the structure of the Excel file to understand how to parse it."""
    print(f"Looking for file: {file_path}")
    print(f"File exists: {os.path.exists(file_path)}")
    
    if not os.path.exists(file_path):
        print(f"File {file_path} not found")
        return
        
    workbook = openpyxl.load_workbook(file_path)
    print(f"Excel file: {file_path}")
    print(f"Sheet names: {workbook.sheetnames}")
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Dimensions: {sheet.dimensions}")
        
        # Print first few rows to understand the structure
        print("First 10 rows:")
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_num > 10:
                break
            print(f"Row {row_num}: {row}")
            
    workbook.close()

if __name__ == "__main__":
    excel_file = r"c:\Dev\Autofire\Database Export.xlsx"
    analyze_excel_structure(excel_file)