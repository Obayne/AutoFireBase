#!/usr/bin/env python3
"""
Simple Excel to Database Importer for AutoFire.

This is a simplified version that uses only openpyxl and standard library modules.
"""

import os
import sys
import sqlite3
import json
import openpyxl
from pathlib import Path

# Add project root to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from db.loader import connect, ensure_schema

def normalize_boolean(value):
    """Convert various boolean representations to Python boolean."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ('true', 'yes', '1', 'y', 't', 'x')
    if isinstance(value, (int, float)):
        return bool(value)
    return False

def parse_candela_options(value):
    """Parse candela options from string to list of integers."""
    if not value:
        return []
    if isinstance(value, (list, tuple)):
        return [int(x) for x in value if str(x).strip().isdigit()]
    if isinstance(value, str):
        # Split by comma and convert to integers
        return [int(x.strip()) for x in str(value).split(',') if x.strip().isdigit()]
    return []

def get_column_index(headers, column_name):
    """Get the index of a column by name (case insensitive)."""
    column_name = column_name.lower()
    for i, header in enumerate(headers):
        if isinstance(header, str) and header.lower() == column_name:
            return i
    return None

def safe_float(value, default=0.0):
    """Safely convert value to float."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

def safe_str(value, default=''):
    """Safely convert value to string."""
    if value is None:
        return default
    return str(value)

def import_excel_to_database(excel_file_path, sheet_name='Database Devices'):
    """
    Import device data from Excel file to database using only openpyxl.
    
    Args:
        excel_file_path (str): Path to the Excel file
        sheet_name (str): Name of the sheet containing device data
    """
    print(f"Importing data from: {excel_file_path}")
    
    # Check if file exists
    if not os.path.exists(excel_file_path):
        print(f"Error: File not found: {excel_file_path}")
        return False
    
    # Read Excel file
    try:
        print("Reading Excel file...")
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
            workbook.close()
            return False
            
        sheet = workbook[sheet_name]
        print(f"Found sheet '{sheet_name}'")
        
        # Read headers
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        print(f"Headers: {headers}")
        
        # Map column indices
        col_indices = {
            'manufacturer': get_column_index(headers, 'Manufacturer'),
            'category': get_column_index(headers, 'Category'),
            'subcategory1': get_column_index(headers, 'SubCategory1'),
            'model': get_column_index(headers, 'Model'),
            'part_number': get_column_index(headers, 'PartNo'),
            'description': get_column_index(headers, 'Description'),
            'symbol': get_column_index(headers, 'PartType'),
            'reqd_standby_current': get_column_index(headers, 'ReqdStandbyCurrent'),
            'reqd_alarm_current': get_column_index(headers, 'ReqdAlarmCurrent'),
            'nominal_voltage': get_column_index(headers, 'NominalVoltage'),
            'min_voltage': get_column_index(headers, 'MinVoltage')
        }
        
        # Show column mapping
        print("Column mapping:")
        for col_name, col_index in col_indices.items():
            if col_index is not None:
                print(f"  {col_name}: column {col_index} ('{headers[col_index]}')")
            else:
                print(f"  {col_name}: not found")
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Connect to database
    try:
        print("Connecting to database...")
        con = connect()
        ensure_schema(con)
        cur = con.cursor()
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return False
    
    # Import data
    imported_count = 0
    error_count = 0
    row_count = 0
    
    try:
        # Process rows (skip header)
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_num == 1:  # Skip header row
                continue
                
            row_count += 1
            if row_count % 100 == 0:
                print(f"Processing row {row_count}...")
                
            try:
                # Extract values with defaults
                def get_cell_value(col_name, default=''):
                    col_idx = col_indices.get(col_name)
                    if col_idx is not None and col_idx < len(row):
                        value = row[col_idx]
                        return value if value is not None else default
                    return default
                
                manufacturer = safe_str(get_cell_value('manufacturer'), '(Unknown)')
                category = safe_str(get_cell_value('category'), 'Fire Alarm')
                subcategory1 = safe_str(get_cell_value('subcategory1'), '')
                model = safe_str(get_cell_value('model'), '')
                part_number = safe_str(get_cell_value('part_number'), '')
                description = safe_str(get_cell_value('description'), '')
                symbol = safe_str(get_cell_value('symbol'), 'DEV')
                reqd_standby_current = safe_float(get_cell_value('reqd_standby_current'), 0.0)
                reqd_alarm_current = safe_float(get_cell_value('reqd_alarm_current'), 0.0)
                nominal_voltage = safe_float(get_cell_value('nominal_voltage'), 24.0)
                min_voltage = safe_float(get_cell_value('min_voltage'), 20.0)
                
                # Handle missing model
                if not model and part_number:
                    model = part_number
                elif not model and not part_number:
                    model = f"{manufacturer}-{symbol}" if manufacturer and symbol else f"MODEL-{row_count}"
                
                # Skip empty rows
                if not any([manufacturer, category, model, description, symbol]):
                    continue
                
                # Insert or get manufacturer ID
                cur.execute("INSERT OR IGNORE INTO manufacturers(name) VALUES(?)", (manufacturer,))
                cur.execute("SELECT id FROM manufacturers WHERE name=?", (manufacturer,))
                manufacturer_row = cur.fetchone()
                if manufacturer_row:
                    manufacturer_id = manufacturer_row[0]
                else:
                    cur.execute("INSERT INTO manufacturers(name) VALUES(?)", (manufacturer,))
                    manufacturer_id = cur.lastrowid
                
                # Determine device type based on category and subcategory
                device_type = 'Detector'
                if 'Control' in category or 'Control' in subcategory1:
                    device_type = 'Control'
                elif 'Notification' in category or 'Notification' in subcategory1:
                    device_type = 'Notification'
                elif 'Initiating' in category or 'Initiating' in subcategory1:
                    device_type = 'Initiating'
                elif 'Speaker' in category or 'Speaker' in subcategory1:
                    device_type = 'Notification'
                elif 'Detector' in category or 'Detector' in subcategory1:
                    device_type = 'Detector'
                
                # Ensure device type exists and get its ID
                cur.execute("INSERT OR IGNORE INTO device_types(code) VALUES(?)", (device_type,))
                cur.execute("SELECT id FROM device_types WHERE code=?", (device_type,))
                type_row = cur.fetchone()
                if type_row:
                    type_id = type_row[0]
                else:
                    # This shouldn't happen, but just in case
                    cur.execute("INSERT INTO device_types(code) VALUES(?)", (device_type,))
                    type_id = cur.lastrowid
                
                # Insert or get system category ID
                cur.execute("INSERT OR IGNORE INTO system_categories(name) VALUES(?)", (category,))
                cur.execute("SELECT id FROM system_categories WHERE name=?", (category,))
                category_row = cur.fetchone()
                category_id = category_row[0] if category_row else None
                
                # Prepare properties JSON
                properties = {
                    'description': description,
                    'reqd_standby_current': reqd_standby_current,
                    'reqd_alarm_current': reqd_alarm_current,
                    'nominal_voltage': nominal_voltage,
                    'min_voltage': min_voltage
                }
                
                # Insert device
                cur.execute("""
                    INSERT INTO devices(manufacturer_id, type_id, category_id, model, name, symbol, properties_json) 
                    VALUES(?,?,?,?,?,?,?)""",
                    (manufacturer_id, type_id, category_id, model, description, symbol, json.dumps(properties))
                )
                device_id = cur.lastrowid
                
                # Handle fire alarm specific specs
                if category == 'Fire Alarm' or 'Fire Alarm' in category:
                    max_current_ma = max(reqd_standby_current * 1000, reqd_alarm_current * 1000)  # Convert to mA
                    
                    cur.execute("""
                        INSERT OR REPLACE INTO fire_alarm_device_specs 
                        (device_id, device_class, max_current_ma, voltage_v, slc_compatible, nac_compatible, addressable, candela_options)
                        VALUES(?,?,?,?,?,?,?,?)""",
                        (device_id, device_type, max_current_ma, nominal_voltage, True, True, True, None)
                    )
                
                imported_count += 1
                if imported_count % 50 == 0:
                    print(f"Imported {imported_count} devices...")
                    
            except Exception as e:
                print(f"Error importing row {row_count}: {e}")
                error_count += 1
                continue
                
    except Exception as e:
        print(f"Error processing Excel data: {e}")
        import traceback
        traceback.print_exc()
        error_count += 1
    finally:
        workbook.close()
    
    # Commit changes
    try:
        con.commit()
        print(f"Database changes committed.")
    except Exception as e:
        print(f"Error committing database changes: {e}")
        error_count += 1
    finally:
        con.close()
    
    print(f"Import completed. Rows processed: {row_count}, Successfully imported: {imported_count}, Errors: {error_count}")
    return error_count == 0

if __name__ == "__main__":
    # Check command line arguments
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
        sheet_name = sys.argv[2] if len(sys.argv) > 2 else 'Database Devices'
        import_excel_to_database(excel_file, sheet_name)
    else:
        print("Usage:")
        print("  python simple_excel_import.py <excel_file> [sheet_name]")
        print("\nTrying default file...")
        
        # Default file
        default_file = r"c:\Dev\Autofire\Database Export.xlsx"
        if os.path.exists(default_file):
            import_excel_to_database(default_file, "Database Devices")
        else:
            print(f"Default file not found: {default_file}")
            print("Please provide the path to your Excel file as a command line argument.")