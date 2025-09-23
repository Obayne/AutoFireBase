import openpyxl
import os

def read_excel_simple():
    file_path = r"c:\Dev\Autofire\Database Export.xlsx"
    print(f"Current working directory: {os.getcwd()}")
    print(f"File exists: {os.path.exists(file_path)}")
    print(f"Full file path: {file_path}")
    
    if not os.path.exists(file_path):
        print("File does not exist!")
        return
        
    try:
        print("Attempting to load workbook...")
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        print("Workbook loaded successfully!")
        print(f"Sheet names: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            print(f"\nProcessing sheet: {sheet_name}")
            
            # Read first few rows
            print("First 5 rows:")
            sheet = workbook[sheet_name]
            for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_num > 5:
                    break
                print(f"  Row {row_num}: {row}")
                
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    read_excel_simple()